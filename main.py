"""
生产人员KPI计算程序 - 入口文件（支持多员工批量处理）

使用方法:
    python main.py

程序将:
1. 读取kpi.xlsx中的所有工作表（每个工作表代表一个员工）
2. 为每个员工计算KPI指标
3. 生成个人详细报告
4. 生成所有员工的汇总对比报告
"""

import os
import sys
from datetime import datetime
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import STANDARD_PARAMS, CHINESE_FONT, COLOR_PRIMARY, COLOR_HEADER, COLOR_GOLD
from kpi_calculator import KPICalculator
from kpi_report_generator import KPIReportGenerator, generate_console_report


class MultiEmployeeKPICalculator:
    """多员工KPI批量计算器"""

    def __init__(self, input_file: str, output_dir: str):
        self.input_file = input_file
        self.output_dir = output_dir
        self.all_reports: List[Dict] = []
        self.errors: List[str] = []

    def process_all_employees(self) -> None:
        """处理所有员工的数据"""
        # 读取所有sheet名称
        xl = pd.ExcelFile(self.input_file)
        sheet_names = xl.sheet_names

        print(f"\n发现 {len(sheet_names)} 个工作表，将逐个处理...")
        print("-" * 60)

        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"\n[{i}/{len(sheet_names)}] 正在处理工作表: {sheet_name}")
            try:
                report = self._process_single_employee(sheet_name)
                if report:
                    self.all_reports.append(report)
                    print(f"  [OK] 员工: {report['employee_info']['employee_name']} "
                          f"(工号: {report['employee_info']['employee_id']}) - "
                          f"得分: {report['total_score']:.2f}分")
            except Exception as e:
                error_msg = f"处理工作表 '{sheet_name}' 时出错: {str(e)}"
                self.errors.append(error_msg)
                print(f"  [ERROR] {error_msg}")

    def _process_single_employee(self, sheet_name: str) -> Dict:
        """处理单个员工的数据"""
        calculator = KPICalculator(standard_params=STANDARD_PARAMS)

        # 读取指定sheet的数据（第0行是公司名称，第1行是员工信息，第2行是表头）
        df = pd.read_excel(self.input_file, sheet_name=sheet_name, header=None)

        # 检查第0行是否为公司名称（特征：包含"公司"字样）
        first_row = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ""
        if "公司" in first_row or "知恩" in first_row or "有限公司" in first_row:
            # 新格式：第0行是公司名称，第1行是员工信息，第2行是表头
            company_name = first_row
            employee_info_row = 1
            data_header_row = 2
        else:
            # 旧格式：第0行是员工信息，第1行是表头
            company_name = ""
            employee_info_row = 0
            data_header_row = 1

        # 提取员工信息
        calculator.employee_id = str(df.iloc[employee_info_row, 1]) if pd.notna(df.iloc[employee_info_row, 1]) else ""
        calculator.employee_name = str(df.iloc[employee_info_row, 4]) if pd.notna(df.iloc[employee_info_row, 4]) else ""

        # 读取数据行（根据表头位置调整）
        data_df = pd.read_excel(self.input_file, sheet_name=sheet_name, header=data_header_row)

        # 标准化列名 (支持设备列的多语言变体)
        # 注意：只映射"设备名称"列，不映射"设备编号"列
        column_mapping = {
            '日期': 'date',
            '工站': 'station',
            '工厂型号': 'product',
            '生產時數': 'hours',
            '良品數量': 'good_qty',
            '返工品數量': 'rework_qty',
            '報廢品數量': 'scrap_qty',
            '设备名称': 'device',
            '設備名稱': 'device',
        }

        actual_columns = {}
        used_targets = set()
        for col in data_df.columns:
            col_str = str(col).strip()
            for key, value in column_mapping.items():
                if key in col_str and value not in used_targets:
                    actual_columns[col] = value
                    used_targets.add(value)
                    break

        data_df = data_df.rename(columns=actual_columns)
        required_cols = ['date', 'station', 'product', 'hours', 'good_qty', 'rework_qty', 'scrap_qty']
        available_cols = [c for c in required_cols if c in data_df.columns]
        calculator.raw_data = data_df[available_cols].copy()

        # 检测设备列，如果不存在则添加并默认为空字符串（手工作业）
        if 'device' in data_df.columns:
            calculator.raw_data['device'] = data_df['device'].fillna('').astype(str).str.strip()
        else:
            calculator.raw_data['device'] = ''

        # 数据类型转换
        calculator.raw_data['hours'] = pd.to_numeric(calculator.raw_data['hours'], errors='coerce').fillna(0)
        calculator.raw_data['good_qty'] = pd.to_numeric(calculator.raw_data['good_qty'], errors='coerce').fillna(0)
        calculator.raw_data['rework_qty'] = pd.to_numeric(calculator.raw_data['rework_qty'], errors='coerce').fillna(0)
        calculator.raw_data['scrap_qty'] = pd.to_numeric(calculator.raw_data['scrap_qty'], errors='coerce').fillna(0)

        # 处理工站名称中包含空格的情况（多工站拆分）
        calculator.raw_data = self._expand_multi_station_rows(calculator.raw_data)

        # 处理工站名称中包含空格的情况（多工站拆分）
        calculator.raw_data = self._expand_multi_station_rows(calculator.raw_data)

        # 提取月份
        if 'date' in calculator.raw_data.columns and len(calculator.raw_data) > 0:
            first_date = pd.to_datetime(calculator.raw_data['date'].iloc[0])
            calculator.month = first_date.strftime('%Y年%m月')

        # 执行计算
        calculator.aggregate_by_product_station()
        calculator.calculate_all_kpis()

        return calculator.generate_report()

    def _expand_multi_station_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        展开包含多个工站的行
        例如: '纸箱打两条                      黄色包装带' -> 拆分为两行
        """
        if 'station' not in df.columns:
            return df

        expanded_rows = []
        for idx, row in df.iterrows():
            station_value = str(row['station']) if pd.notna(row['station']) else ''

            # 按空白字符分割工站名称
            stations = [s.strip() for s in station_value.split() if s.strip()]

            if len(stations) <= 1:
                # 单行工站，直接保留
                expanded_rows.append(row.to_dict())
            else:
                # 多工站行，拆分为多行，数据按比例分摊
                n_stations = len(stations)
                for station in stations:
                    new_row = row.to_dict().copy()
                    new_row['station'] = station
                    # 工时和产量按工站数量平均分摊
                    new_row['hours'] = row['hours'] / n_stations
                    new_row['good_qty'] = int(row['good_qty'] / n_stations)
                    new_row['rework_qty'] = int(row['rework_qty'] / n_stations)
                    new_row['scrap_qty'] = int(row['scrap_qty'] / n_stations)
                    expanded_rows.append(new_row)

        return pd.DataFrame(expanded_rows)

    def generate_individual_reports(self) -> None:
        """为每个员工生成个人报告"""
        print("\n" + "=" * 60)
        print("正在生成个人详细报告...")
        print("=" * 60)

        for report in self.all_reports:
            emp_name = report['employee_info']['employee_name']
            emp_id = report['employee_info']['employee_id']
            month = report['employee_info']['month'].replace('年', '').replace('月', '')

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(
                self.output_dir,
                f"KPI报告_{emp_name}_{emp_id}_{month}_{timestamp}.xlsx"
            )

            report_generator = KPIReportGenerator(report)
            report_generator.generate_excel(output_file)

    def generate_summary_report(self) -> str:
        """生成所有员工的汇总对比报告"""
        print("\n" + "=" * 60)
        print("正在生成员工汇总对比报告...")
        print("=" * 60)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = os.path.join(self.output_dir, f"KPI汇总报告_{timestamp}.xlsx")

        wb = Workbook()
        wb.remove(wb.active)

        # 创建汇总表
        self._add_summary_comparison_sheet(wb)

        # 创建排名表
        self._add_ranking_sheet(wb)

        wb.save(summary_file)
        print(f"汇总报告已生成: {summary_file}")
        return summary_file

    def _add_summary_comparison_sheet(self, wb):
        """添加汇总对比表"""
        ws = wb.create_sheet('员工KPI汇总对比', 0)

        # 标题
        ws.merge_cells('A1:K1')
        cell = ws.cell(1, 1, '生产人员KPI绩效汇总对比表')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=16, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
        ws.row_dimensions[1].height = 30

        # 表头
        headers = ['排名', '工号', '姓名', '月份', '工时达成率', '良品达成率', '人时产出达成率',
                   '返工率控制', '报废率控制', '综合得分', '等级']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

        # 按得分排序
        sorted_reports = sorted(self.all_reports, key=lambda x: x['total_score'], reverse=True)

        # 数据行
        for i, report in enumerate(sorted_reports, 1):
            data_row = row + i
            emp_info = report['employee_info']
            kpi_results = {r['indicator_name']: r for r in report['kpi_results']}

            # 排名
            cell = ws.cell(data_row, 1, i)
            cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            if i == 1:
                cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            elif i == 2:
                cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            elif i == 3:
                cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')

            # 基本信息
            ws.cell(data_row, 2, emp_info.get('employee_id', ''))
            ws.cell(data_row, 3, emp_info.get('employee_name', ''))
            ws.cell(data_row, 4, emp_info.get('month', ''))

            # KPI指标
            ws.cell(data_row, 5, f"{kpi_results.get('工时达成率', {}).get('actual_value', 0):.2f}%")
            ws.cell(data_row, 6, f"{kpi_results.get('良品达成率', {}).get('actual_value', 0):.2f}%")
            ws.cell(data_row, 7, f"{kpi_results.get('人时产出达成率', {}).get('actual_value', 0):.2f}%")
            ws.cell(data_row, 8, f"{kpi_results.get('返工率控制', {}).get('actual_value', 0):.2f}%")
            ws.cell(data_row, 9, f"{kpi_results.get('报废率控制', {}).get('actual_value', 0):.2f}%")

            # 综合得分
            score_cell = ws.cell(data_row, 10, f"{report['total_score']:.2f}")
            score_cell.font = Font(name=CHINESE_FONT, bold=True, size=11)

            # 等级
            grade_cell = ws.cell(data_row, 11, report['total_grade'])
            grade_cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            if report['total_grade'] == '甲':
                grade_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                grade_cell.font = Font(name=CHINESE_FONT, bold=True, size=11, color='FFFFFF')
            elif report['total_grade'] == '丁':
                grade_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                grade_cell.font = Font(name=CHINESE_FONT, bold=True, size=11, color='FFFFFF')

            # 设置边框和对齐
            for col in range(1, 12):
                cell = ws.cell(data_row, col)
                cell.font = Font(name=CHINESE_FONT, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )

            # 交替行背景
            if i % 2 == 0:
                for col in range(1, 12):
                    ws.cell(data_row, col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        for col in ['E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 14
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 10

    def _add_ranking_sheet(self, wb):
        """添加排名分析表 - 专业版"""
        from openpyxl.chart import BarChart, Reference

        ws = wb.create_sheet('绩效排名分析')

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # 主标题
        ws.merge_cells('A1:G1')
        cell = ws.cell(1, 1, '生产人员KPI绩效深度分析报告')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=16, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
        ws.row_dimensions[1].height = 35

        # ===== 第一部分：等级分布统计 =====
        row = 3
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row, 1, '一、绩效等级分布')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=13, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 25
        row += 1

        # 表头
        headers = ['等级', '人数', '占比', '人员清单', '最低分', '最高分', '平均分']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 25
        row += 1

        # 按等级统计
        grade_data = {}
        for report in self.all_reports:
            grade = report['total_grade']
            if grade not in grade_data:
                grade_data[grade] = {'names': [], 'scores': []}
            grade_data[grade]['names'].append(report['employee_info']['employee_name'])
            grade_data[grade]['scores'].append(report['total_score'])

        for grade in ['甲', '乙', '丙', '丁']:
            data = grade_data.get(grade, {'names': [], 'scores': []})
            count = len(data['names'])
            percentage = (count / len(self.all_reports) * 100) if self.all_reports else 0
            min_score = min(data['scores']) if data['scores'] else 0
            max_score = max(data['scores']) if data['scores'] else 0
            avg_score = sum(data['scores']) / len(data['scores']) if data['scores'] else 0

            ws.cell(row, 1, f'{grade}等')
            ws.cell(row, 2, count)
            ws.cell(row, 3, f'{percentage:.1f}%')
            ws.cell(row, 4, '、'.join(data['names']) if data['names'] else '-')
            ws.cell(row, 5, f'{min_score:.2f}')
            ws.cell(row, 6, f'{max_score:.2f}')
            ws.cell(row, 7, f'{avg_score:.2f}')

            # 设置格式
            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.font = Font(name=CHINESE_FONT, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

                # 等级颜色标记
                if col == 1:
                    if grade == '甲':
                        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=11, color='FFFFFF')
                    elif grade == '丁':
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=11, color='FFFFFF')

            ws.row_dimensions[row].height = 30
            row += 1

        # ===== 第二部分：各项指标团队分析 =====
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row, 1, '二、KPI指标团队分析')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=13, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 25
        row += 1

        # 表头
        headers = ['KPI指标', '团队平均', '目标/标准', '达成情况', '最高', '最低', '标准差']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 25
        row += 1

        # 统计各指标
        indicators = [
            ('工时达成率', [r for r in self.all_reports], '≥100%'),
            ('良品达成率', [r for r in self.all_reports], '≥100%'),
            ('人时产出达成率', [r for r in self.all_reports], '≥100%'),
            ('返工率控制', [r for r in self.all_reports], '≤0.5%'),
            ('报废率控制', [r for r in self.all_reports], '≤0.1%'),
        ]

        for indicator_name, reports, standard in indicators:
            values = []
            for report in reports:
                for kpi in report['kpi_results']:
                    if kpi['indicator_name'] == indicator_name:
                        values.append(kpi['actual_value'])
                        break

            if values:
                avg_val = sum(values) / len(values)
                max_val = max(values)
                min_val = min(values)
                std_val = (sum((x - avg_val) ** 2 for x in values) / len(values)) ** 0.5

                # 达成情况判断
                if '控制' in indicator_name:
                    status = '达标' if avg_val <= float(standard.replace('≤', '').replace('%', '')) else '超标'
                else:
                    status = '达标' if avg_val >= float(standard.replace('≥', '').replace('%', '')) else '未达标'

                ws.cell(row, 1, indicator_name)
                ws.cell(row, 2, f'{avg_val:.2f}%')
                ws.cell(row, 3, standard)
                ws.cell(row, 4, status)
                ws.cell(row, 5, f'{max_val:.2f}%')
                ws.cell(row, 6, f'{min_val:.2f}%')
                ws.cell(row, 7, f'{std_val:.2f}')

                for col in range(1, 8):
                    cell = ws.cell(row, col)
                    cell.font = Font(name=CHINESE_FONT, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                    # 达成情况颜色标记
                    if col == 4:
                        if status == '达标':
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(name=CHINESE_FONT, size=10, color='006100')
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(name=CHINESE_FONT, size=10, color='9C0006')

                ws.row_dimensions[row].height = 22
                row += 1

        # ===== 第三部分：综合得分统计 =====
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row, 1, '三、综合得分统计')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=13, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 25
        row += 1

        scores = [r['total_score'] for r in self.all_reports]

        # 统计表
        stats_data = [
            ('最高分', f'{max(scores):.2f}分', self.all_reports[scores.index(max(scores))]['employee_info']['employee_name'] if scores else '-'),
            ('最低分', f'{min(scores):.2f}分', self.all_reports[scores.index(min(scores))]['employee_info']['employee_name'] if scores else '-'),
            ('平均分', f'{sum(scores)/len(scores):.2f}分', '-'),
            ('中位数', f'{sorted(scores)[len(scores)//2]:.2f}分' if scores else '-', '-'),
            ('标准差', f'{(sum((x - sum(scores)/len(scores))**2 for x in scores)/len(scores))**0.5:.2f}', '团队离散程度'),
        ]

        for label, value, note in stats_data:
            ws.cell(row, 1, label)
            ws.cell(row, 2, value)
            ws.merge_cells(f'C{row}:G{row}')
            ws.cell(row, 3, note)

            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.font = Font(name=CHINESE_FONT, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if col == 1:
                    cell.font = Font(name=CHINESE_FONT, bold=True, size=10)
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')

            ws.row_dimensions[row].height = 22
            row += 1

        # ===== 第四部分：排名明细 =====
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row, 1, '四、绩效排名明细')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=13, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 25
        row += 1

        # 排名表头
        headers = ['排名', '工号', '姓名', '综合得分', '等级', '相对排名', '差距分析']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 25
        row += 1

        # 排序后的报告
        sorted_reports = sorted(self.all_reports, key=lambda x: x['total_score'], reverse=True)
        avg_score = sum(scores) / len(scores) if scores else 0

        for i, report in enumerate(sorted_reports, 1):
            emp_info = report['employee_info']
            score = report['total_score']
            grade = report['total_grade']

            ws.cell(row, 1, i)
            ws.cell(row, 2, emp_info.get('employee_id', ''))
            ws.cell(row, 3, emp_info.get('employee_name', ''))
            ws.cell(row, 4, f'{score:.2f}')
            ws.cell(row, 5, grade)

            # 相对排名百分比
            percentile = (len(sorted_reports) - i + 1) / len(sorted_reports) * 100
            ws.cell(row, 6, f'前{percentile:.0f}%')

            # 差距分析
            diff = score - avg_score
            if diff > 0:
                gap_text = f'+{diff:.2f}分 (高于平均)'
            elif diff < 0:
                gap_text = f'{diff:.2f}分 (低于平均)'
            else:
                gap_text = '持平'
            ws.merge_cells(f'G{row}:G{row}')
            ws.cell(row, 7, gap_text)

            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.font = Font(name=CHINESE_FONT, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 排名颜色标记
                if col == 1:
                    if i == 1:
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
                    elif i == 2:
                        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=11)
                    elif i == 3:
                        cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=11, color='FFFFFF')

                # 等级颜色
                if col == 5:
                    if grade == '甲':
                        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=10, color='FFFFFF')
                    elif grade == '丁':
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=10, color='FFFFFF')

                # 差距分析颜色
                if col == 7:
                    if diff > 0:
                        cell.font = Font(name=CHINESE_FONT, size=10, color='006100')
                    elif diff < 0:
                        cell.font = Font(name=CHINESE_FONT, size=10, color='9C0006')

            ws.row_dimensions[row].height = 22
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25


def main():
    """主程序入口"""
    # 设置控制台编码支持中文显示
    import sys
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    print("=" * 70)
    print("生产人员KPI计算程序 - 批量处理版")
    print("=" * 70)

    # 文件路径
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 检查命令行参数
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        if not os.path.isabs(input_file):
            input_file = os.path.join(base_dir, input_file)
    else:
        # 默认文件
        input_file = os.path.join(base_dir, 'kpi.xlsx')

    output_dir = os.path.join(base_dir, 'output')

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件 {input_file}")
        print("请确保kpi.xlsx文件与程序在同一目录下")
        sys.exit(1)

    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"已创建输出目录: {output_dir}")

    print(f"\n输入文件: {input_file}")

    # 初始化批量处理器
    processor = MultiEmployeeKPICalculator(input_file, output_dir)

    # 处理所有员工
    processor.process_all_employees()

    # 生成个人报告
    if processor.all_reports:
        processor.generate_individual_reports()

        # 如果有多个员工，生成汇总报告
        if len(processor.all_reports) > 1:
            processor.generate_summary_report()

        # 在控制台输出汇总
        print("\n" + "=" * 70)
        print("所有员工KPI计算完成！")
        print("=" * 70)
        print(f"\n成功处理: {len(processor.all_reports)} 名员工")

        if processor.errors:
            print(f"处理失败: {len(processor.errors)} 个工作表")
            for error in processor.errors:
                print(f"  - {error}")

        print(f"\n输出目录: {output_dir}")
        print("\n生成的文件包括:")
        print("  1. 个人详细报告: KPI报告_姓名_工号_月份_时间戳.xlsx")
        if len(processor.all_reports) > 1:
            print("  2. 汇总对比报告: KPI汇总报告_时间戳.xlsx")
    else:
        print("\n错误: 未能成功处理任何员工数据")
        if processor.errors:
            print("错误详情:")
            for error in processor.errors:
                print(f"  - {error}")
        sys.exit(1)


if __name__ == '__main__':
    main()
