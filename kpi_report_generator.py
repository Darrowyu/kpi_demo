"""
KPI报告生成器 - 将计算结果导出为Excel文件
使用微软雅黑字体，提供专业的KPI分析报告
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple
from datetime import datetime
import os


# 定义中文字体
CHINESE_FONT = '微软雅黑'
HEADER_FONT_SIZE = 11
NORMAL_FONT_SIZE = 10
TITLE_FONT_SIZE = 16
SUBTITLE_FONT_SIZE = 12

# 定义颜色主题
COLOR_PRIMARY = '4472C4'      # 主色调-蓝色
COLOR_SECONDARY = '5B9BD5'    # 次要色-浅蓝
COLOR_HEADER = 'D9E1F2'       # 表头背景
COLOR_ALT_ROW = 'F2F2F2'      # 交替行背景
COLOR_GOLD = 'FFD700'         # 金色高亮
COLOR_RED = 'FF6B6B'          # 红色警示
COLOR_GREEN = '70AD47'        # 绿色优秀
COLOR_YELLOW = 'FFC000'       # 黄色警告

# 边框样式
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


class KPIReportGenerator:
    """KPI详细报告Excel生成器"""

    def __init__(self, report_data: Dict):
        self.report_data = report_data
        self.employee_info = report_data.get('employee_info', {})
        self.raw_data = report_data.get('raw_data', [])
        self.aggregated_data = report_data.get('aggregated_data', [])
        self.kpi_results = report_data.get('kpi_results', [])
        self.total_score = report_data.get('total_score', 0)
        self.total_grade = report_data.get('total_grade', '丁')

    def _set_cell_font(self, cell, bold=False, size=NORMAL_FONT_SIZE, color='000000'):
        """设置单元格中文字体"""
        cell.font = Font(name=CHINESE_FONT, bold=bold, size=size, color=color)

    def _create_header(self, ws, title: str, start_row: int = 1, end_col: str = 'K') -> int:
        """创建报告标题头"""
        ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
        cell = ws.cell(start_row, 1, title)
        self._set_cell_font(cell, bold=True, size=SUBTITLE_FONT_SIZE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=SUBTITLE_FONT_SIZE, color='FFFFFF')
        return start_row + 2

    def _analyze_kpi_performance(self) -> Dict:
        """分析KPI表现，生成诊断说明和建议"""
        analysis = {
            'strengths': [],      # 优势项
            'weaknesses': [],     # 劣势项
            'risks': [],          # 风险项
            'suggestions': []     # 改进建议
        }

        for kpi in self.kpi_results:
            name = kpi.get('indicator_name', '')
            value = kpi.get('actual_value', 0)
            grade = kpi.get('grade', '')

            if name == '工时达成率':
                if grade == '甲':
                    analysis['strengths'].append('出勤饱满，工时投入充分')
                elif value >= 110:
                    analysis['strengths'].append(f'工时达成率{value:.1f}%，加班较多，工作投入度高')

            elif name == '良品达成率':
                if grade == '甲':
                    analysis['strengths'].append('良品率表现优秀，质量控制好')
                elif grade == '乙':
                    analysis['suggestions'].append('良品率接近标准，需关注返工工序对整体的影响')
                elif grade in ['丙', '丁']:
                    analysis['weaknesses'].append(f'良品率{value:.1f}%，低于标准要求，需加强质量控制')

            elif name == '人时产出达成率':
                if grade == '甲':
                    analysis['strengths'].append('生产效率优秀，人时产出达标')
                elif grade == '丁':
                    analysis['weaknesses'].append(f'人时产出达成率仅{value:.1f}%，效率偏低，需分析具体原因')
                    analysis['suggestions'].append('建议复核标准工时的合理性，特别是盖箱号、N95装箱等工序')
                    analysis['risks'].append('部分工序效率严重偏低，可能影响整体交付')

            elif name == '返工率控制':
                if grade == '甲':
                    analysis['strengths'].append('返工率控制良好，一次合格率达标')
                elif grade == '丁':
                    analysis['weaknesses'].append(f'返工率{value:.2f}%，远超标准上限(0.5%)')
                    analysis['risks'].append('返工率高企，造成工时浪费和成本增加')
                    analysis['suggestions'].append('紧急排查返工原因，加强首件检验和过程控制')

            elif name == '报废率控制':
                if grade == '甲':
                    analysis['strengths'].append('无报废品，成本控制良好')
                elif grade in ['丙', '丁']:
                    analysis['risks'].append('报废率偏高，造成物料损失')

        return analysis

    def _add_summary_sheet(self, wb):
        """添加最终绩效汇总表 - 增强版，包含详细诊断和建议"""
        ws = wb.create_sheet('最终绩效汇总', 0)

        # 主标题
        emp_name = self.employee_info.get('employee_name', '')
        emp_id = self.employee_info.get('employee_id', '')
        month = self.employee_info.get('month', '')

        ws.merge_cells('A1:H1')
        cell = ws.cell(1, 1, f'生产人员月度KPI绩效评估报告')
        cell.font = Font(name=CHINESE_FONT, bold=True, size=TITLE_FONT_SIZE, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
        ws.row_dimensions[1].height = 30

        # 员工基本信息
        ws.merge_cells('A2:H2')
        info_text = f"员工姓名：{emp_name}    工号：{emp_id}    评估月份：{month}"
        cell = ws.cell(2, 1, info_text)
        self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        ws.row_dimensions[2].height = 25

        # 第一部分：KPI指标汇总表
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '一、KPI指标达成情况')
        self._set_cell_font(cell, bold=True, size=SUBTITLE_FONT_SIZE, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 22

        # 表头
        row = 5
        headers = ['KPI指标', '实际达成值', '目标/标准', '等级', '原始分', '权重', '加权得分', '诊断说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row].height = 30

        # KPI数据行
        diagnostics = {
            '工时达成率': lambda g, v: '出勤饱满' if g == '甲' else ('加班较多' if v >= 110 else '出勤不足'),
            '良品达成率': lambda g, v: '质量优秀' if g == '甲' else ('接近标准' if g == '乙' else '需改进'),
            '人时产出达成率': lambda g, v: '效率达标' if g == '甲' else ('效率偏低' if g == '丁' else '基本达标'),
            '返工率控制': lambda g, v: '控制良好' if g == '甲' else ('超标需整改' if g == '丁' else '需关注'),
            '报废率控制': lambda g, v: '零报废' if v == 0 else ('控制良好' if g == '甲' else '需改进')
        }

        standards = {
            '工时达成率': '≥100%',
            '良品达成率': '≥100%',
            '人时产出达成率': '≥100%',
            '返工率控制': '≤0.5%',
            '报废率控制': '≤0.1%'
        }

        for i, kpi in enumerate(self.kpi_results, 6):
            indicator = kpi.get('indicator_name', '')
            value = kpi.get('actual_value', 0)
            grade = kpi.get('grade', '')
            raw = kpi.get('raw_score', 0)
            weight = kpi.get('weight', 0)
            weighted = kpi.get('weighted_score', 0)

            # 格式化数值
            value_str = f"{value:.2f}%"
            standard_str = standards.get(indicator, '')

            # 根据等级设置背景色
            if grade == '甲':
                bg_color = COLOR_GREEN
            elif grade == '丁':
                bg_color = COLOR_RED
            else:
                bg_color = None

            cells_data = [
                (indicator, False),
                (value_str, False),
                (standard_str, False),
                (grade, True),
                (raw, False),
                (f"{weight*100:.0f}%", False),
                (f"{weighted:.2f}", False),
                (diagnostics.get(indicator, lambda g, v: '')(grade, value), False)
            ]

            for col, (val, is_grade) in enumerate(cells_data, 1):
                cell = ws.cell(i, col, val)
                self._set_cell_font(cell, bold=is_grade, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if bg_color and col in [4]:  # 等级列高亮
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    if bg_color == COLOR_GREEN:
                        cell.font = Font(name=CHINESE_FONT, bold=True, size=NORMAL_FONT_SIZE, color='FFFFFF')

            ws.row_dimensions[i].height = 25

        # 合计行
        total_row = 6 + len(self.kpi_results)
        ws.cell(total_row, 1, '月度综合绩效得分')
        ws.merge_cells(f'D{total_row}:F{total_row}')
        ws.cell(total_row, 4, f'{self.total_grade}等')
        ws.cell(total_row, 7, f"{self.total_score:.2f}分")

        # 合计行格式
        for col in range(1, 8):
            cell = ws.cell(total_row, col)
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        ws.row_dimensions[total_row].height = 28

        # 第二部分：绩效分析
        row = total_row + 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '二、绩效表现分析')
        self._set_cell_font(cell, bold=True, size=SUBTITLE_FONT_SIZE, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 22

        analysis = self._analyze_kpi_performance()

        # 优势项
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '【表现优势】')
        self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE, color=COLOR_GREEN)

        if analysis['strengths']:
            for i, strength in enumerate(analysis['strengths'], 1):
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row, 1, f"  {i}. {strength}")
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[row].height = 22
        else:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row, 1, "  暂无突出优势项，需全面提升")
            self._set_cell_font(cell, size=NORMAL_FONT_SIZE, color='999999')

        # 劣势项
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '【待改进项】')
        self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE, color=COLOR_RED)

        if analysis['weaknesses']:
            for i, weakness in enumerate(analysis['weaknesses'], 1):
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row, 1, f"  {i}. {weakness}")
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[row].height = 22
        else:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row, 1, "  无明显劣势项，保持当前水平")
            self._set_cell_font(cell, size=NORMAL_FONT_SIZE, color='999999')

        # 风险提示
        if analysis['risks']:
            row += 2
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row, 1, '【风险提示】')
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE, color=COLOR_YELLOW)

            for i, risk in enumerate(analysis['risks'], 1):
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row, 1, f"  ⚠ {risk}")
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE, color='CC6600')
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[row].height = 22

        # 改进建议
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '【改进建议】')
        self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE, color=COLOR_SECONDARY)

        if analysis['suggestions']:
            for i, suggestion in enumerate(analysis['suggestions'], 1):
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row, 1, f"  {i}. {suggestion}")
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[row].height = 22
        else:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row, 1, "  继续保持当前工作状态，争取更高绩效")
            self._set_cell_font(cell, size=NORMAL_FONT_SIZE, color='999999')

        # 第三部分：数据统计
        row += 3
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '三、生产数据统计')
        self._set_cell_font(cell, bold=True, size=SUBTITLE_FONT_SIZE, color=COLOR_PRIMARY)
        ws.row_dimensions[row].height = 22

        # 统计卡片
        total_hours = sum(a.get('total_hours', 0) for a in self.aggregated_data)
        total_good = sum(a.get('total_good', 0) for a in self.aggregated_data)
        total_rework = sum(a.get('total_rework', 0) for a in self.aggregated_data)
        station_count = len(self.aggregated_data)

        stats = [
            (f'总工时', f'{total_hours:.1f}小时', '应出勤176小时'),
            (f'总产量', f'{total_good:,}件', '良品总数'),
            (f'返工数', f'{total_rework:,}件', '需返工处理'),
            (f'作业范围', f'{station_count}个工序', '涉及产品+工站')
        ]

        row += 1
        for i, (label, value, note) in enumerate(stats):
            col_start = i * 2 + 1
            col_end = col_start + 1
            ws.merge_cells(f'{chr(64+col_start)}{row}:{chr(64+col_end)}{row}')
            cell = ws.cell(row, col_start, f"{label}：{value}\n({note})")
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.border = thin_border

        ws.row_dimensions[row].height = 40

        # 底部说明
        row += 3
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row, 1, '评估说明：')
        self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE, color='666666')

        explanations = [
            '1. 本报告基于实际生产数据计算，确保客观公正；',
            '2. KPI指标权重：工时达成率10%、良品达成率30%、人时产出达成率30%、返工率控制15%、报废率控制15%；',
            '3. 等级标准：甲等(10分)、乙等(8分)、丙等(6分)、丁等(4分)；',
            '4. 综合得分6.0-7.99分为乙等，该员工本月绩效为乙等（偏下）。'
        ]

        for explanation in explanations:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row, 1, f"   {explanation}")
            self._set_cell_font(cell, size=NORMAL_FONT_SIZE-1, color='888888')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 调整列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 35

    def _add_raw_data_sheet(self, wb):
        """添加原始生产数据明细表"""
        ws = wb.create_sheet('原始生产数据明细')

        # 标题
        row = self._create_header(ws, '第一部分：原始生产数据明细', 1, 'J')

        # 表头
        headers = ['序号', '日期', '产品', '工站', '工时(h)', '良品数', '返工数', '报废数', '实际人时产出', '实际良品率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        for i, record in enumerate(self.raw_data, 1):
            hours = record.get('hours', 0)
            good = record.get('good_qty', 0)
            rework = record.get('rework_qty', 0)
            scrap = record.get('scrap_qty', 0)
            total = good + rework + scrap

            output = good / hours if hours > 0 else 0
            quality_rate = good / total if total > 0 else 0

            data_row = row + i
            ws.cell(data_row, 1, i)
            ws.cell(data_row, 2, str(record.get('date', '')))
            ws.cell(data_row, 3, record.get('product', ''))
            ws.cell(data_row, 4, record.get('station', ''))
            ws.cell(data_row, 5, hours)
            ws.cell(data_row, 6, good)
            ws.cell(data_row, 7, rework)
            ws.cell(data_row, 8, scrap)
            ws.cell(data_row, 9, f"{output:.0f}件/h")
            ws.cell(data_row, 10, f"{quality_rate*100:.1f}%")

            # 设置字体和对齐
            for col in range(1, 11):
                cell = ws.cell(data_row, col)
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 交替行背景
            if i % 2 == 0:
                for col in range(1, 11):
                    ws.cell(data_row, col).fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

        # 合计行
        total_row = row + len(self.raw_data) + 1
        total_hours = sum(r.get('hours', 0) for r in self.raw_data)
        total_good = sum(r.get('good_qty', 0) for r in self.raw_data)
        total_rework = sum(r.get('rework_qty', 0) for r in self.raw_data)
        total_scrap = sum(r.get('scrap_qty', 0) for r in self.raw_data)

        ws.cell(total_row, 1, '合计')
        ws.cell(total_row, 5, total_hours)
        ws.cell(total_row, 6, total_good)
        ws.cell(total_row, 7, total_rework)
        ws.cell(total_row, 8, total_scrap)

        for col in range(1, 11):
            cell = ws.cell(total_row, col)
            self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 调整列宽
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 14

    def _add_aggregated_sheet(self, wb):
        """添加中间计算表（按产品+工站聚合）"""
        ws = wb.create_sheet('中间计算表-产品工站聚合')

        # 标题
        row = self._create_header(ws, '第二部分：按"产品+设备+工站"维度聚合计算（中间计算表）', 1, 'K')

        # 表头
        headers = ['产品型号', '设备名称', '工站', '总工时(h)', '总良品数', '总返工数', '总报废数',
                   '实际人时产出', '实际良品率', '工时占比', '产量占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        for i, agg in enumerate(self.aggregated_data, 1):
            data_row = row + i
            device_name = agg.device if hasattr(agg, 'device') else agg.get('device', '')
            # 设备为空时显示"手工"（处理None、nan和空字符串）
            import pandas as pd
            device_display = device_name if pd.notna(device_name) and str(device_name).strip() else '手工'

            ws.cell(data_row, 1, agg.product if hasattr(agg, 'product') else agg.get('product', ''))
            ws.cell(data_row, 2, device_display)
            ws.cell(data_row, 3, agg.station if hasattr(agg, 'station') else agg.get('station', ''))
            ws.cell(data_row, 4, agg.total_hours if hasattr(agg, 'total_hours') else agg.get('total_hours', 0))
            ws.cell(data_row, 5, agg.total_good if hasattr(agg, 'total_good') else agg.get('total_good', 0))
            ws.cell(data_row, 6, agg.total_rework if hasattr(agg, 'total_rework') else agg.get('total_rework', 0))
            ws.cell(data_row, 7, agg.total_scrap if hasattr(agg, 'total_scrap') else agg.get('total_scrap', 0))
            ws.cell(data_row, 8, f"{agg.actual_output if hasattr(agg, 'actual_output') else agg.get('actual_output', 0):.0f}件/h")
            ws.cell(data_row, 9, f"{agg.actual_quality_rate if hasattr(agg, 'actual_quality_rate') else agg.get('actual_quality_rate', 0)*100:.1f}%")
            ws.cell(data_row, 10, f"{agg.hours_ratio if hasattr(agg, 'hours_ratio') else agg.get('hours_ratio', 0)*100:.2f}%")
            ws.cell(data_row, 11, f"{agg.output_ratio if hasattr(agg, 'output_ratio') else agg.get('output_ratio', 0)*100:.2f}%")

            # 设置字体和对齐
            for col in range(1, 11):
                cell = ws.cell(data_row, col)
                self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 交替行背景
            if i % 2 == 0:
                for col in range(1, 11):
                    ws.cell(data_row, col).fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

        # 合计行
        total_row = row + len(self.aggregated_data) + 1
        total_hours = sum(a.get('total_hours', 0) for a in self.aggregated_data)
        total_good = sum(a.get('total_good', 0) for a in self.aggregated_data)
        total_rework = sum(a.get('total_rework', 0) for a in self.aggregated_data)
        total_scrap = sum(a.get('total_scrap', 0) for a in self.aggregated_data)

        ws.cell(total_row, 1, '合计/平均')
        ws.cell(total_row, 3, total_hours)
        ws.cell(total_row, 4, total_good)
        ws.cell(total_row, 5, total_rework)
        ws.cell(total_row, 6, total_scrap)
        ws.cell(total_row, 9, '100.00%')
        ws.cell(total_row, 10, '100.00%')

        for col in range(1, 11):
            cell = ws.cell(total_row, col)
            self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
            cell.fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 调整列宽
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 16

    def _add_kpi_detail_sheet(self, wb):
        """添加KPI分指标详细计算过程"""
        ws = wb.create_sheet('KPI分指标详细计算')

        # 标题
        row = self._create_header(ws, '第三部分：KPI分指标详细计算过程', 1, 'G')

        current_row = row

        # 遍历每个KPI指标
        for kpi in self.kpi_results:
            indicator_name = kpi.get('indicator_name', '')
            actual_value = kpi.get('actual_value', 0)
            grade = kpi.get('grade', '')
            raw_score = kpi.get('raw_score', 0)
            weight = kpi.get('weight', 0)
            weighted_score = kpi.get('weighted_score', 0)

            # KPI标题
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws.cell(current_row, 1, f"【{indicator_name}】（权重{weight*100:.0f}%）")
            self._set_cell_font(cell, bold=True, size=HEADER_FONT_SIZE, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 计算结果摘要
            value_str = f"{actual_value:.2f}%"

            # 根据指标类型添加计算说明
            if indicator_name == '工时达成率':
                ws.merge_cells(f'A{current_row}:G{current_row}')
                calc_explain = f'计算公式：工时达成率 = 实际总工时 ÷ 标准工时(176小时) × 100% = {actual_value:.2f}%'
                cell = ws.cell(current_row, 1, calc_explain)
                self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE, color=COLOR_SECONDARY)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                # 添加详细数据行 - 调整为7列
                total_hours = sum(agg.get('total_hours', 0) for agg in self.aggregated_data)
                ws.cell(current_row, 1, '实际总工时：')
                ws.cell(current_row, 2, f'{total_hours:.1f}小时')
                ws.cell(current_row, 3, '标准工时：')
                ws.cell(current_row, 4, '176小时(22天×8小时)')
                ws.cell(current_row, 5, f'{total_hours:.1f}÷176×100%={actual_value:.2f}%')
                ws.cell(current_row, 6, f'等级：{grade}')
                ws.cell(current_row, 7, f'得分：{weighted_score:.2f}分')

                for col in range(1, 8):
                    cell = ws.cell(current_row, col)
                    self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 2

            summary_data = [
                ('实际达成值：', value_str, '等级：', grade, '原始分：', raw_score, f'加权得分：{weighted_score:.2f}分')
            ]

            for data in summary_data:
                for col, val in enumerate(data, 1):
                    cell = ws.cell(current_row, col, val)
                    if col % 2 == 0:  # 数值列
                        self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE, color=COLOR_PRIMARY)
                    else:
                        self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if indicator_name != '工时达成率':
                current_row += 2
            else:
                current_row += 1

            # 如果是良品达成率或人时产出达成率，显示各组合的详细计算
            if indicator_name in ['良品达成率', '人时产出达成率']:
                # 添加公式说明行
                ws.merge_cells(f'A{current_row}:G{current_row}')
                formula_text = '计算公式：加权贡献 = (实际值÷标准值) × 工时占比'
                cell = ws.cell(current_row, 1, formula_text)
                self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE, color=COLOR_SECONDARY)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                # 表头
                headers = ['产品+工站', '实际值', '标准值', '达成率', '工时占比', '加权贡献', '计算过程详细说明']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(current_row, col, header)
                    self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 30
                current_row += 1

                for agg in self.aggregated_data:
                    product = agg.get('product', '')
                    device = agg.get('device', '')
                    station = agg.get('station', '')
                    key = (product, device, station)

                    # 设备为空时显示"手工"
                    device_display = device if device.strip() else '手工'
                    product_station = f"{product}-{device_display}-{station}"

                    # 查询标准参数
                    from config import STANDARD_PARAMS
                    std_config = None
                    # 尝试三维查询(产品, 设备, 工站)
                    three_key = (product, device, station)
                    # 回退到二维查询(产品, '', 工站)
                    fallback_key = (product, '', station)

                    if three_key in STANDARD_PARAMS:
                        std_config = STANDARD_PARAMS[three_key]
                    elif fallback_key in STANDARD_PARAMS:
                        std_config = STANDARD_PARAMS[fallback_key]
                    else:
                        continue  # 跳过无标准参数的组合

                    if indicator_name == '良品达成率':
                        actual = agg.get('actual_quality_rate', 0) * 100
                        std_value = std_config.get('standard_quality_rate', 0.985) * 100
                        achievement = (actual / std_value) if std_value > 0 else 0
                        unit = '%'
                    else:  # 人时产出达成率
                        actual = agg.get('actual_output', 0)
                        std_value = std_config.get('standard_output', 1)
                        achievement = (actual / std_value) if std_value > 0 else 0
                        unit = '件/h'

                    hours_ratio = agg.get('hours_ratio', 0)
                    contribution = achievement * hours_ratio * 100

                    # 计算过程说明
                    calc_process = (f'步骤1: {actual:.2f}{unit} ÷ {std_value}{unit} = {achievement*100:.2f}%\n'
                                   f'步骤2: {achievement*100:.2f}% × {hours_ratio*100:.2f}% = {contribution:.2f}%')

                    data_cells = [
                        product_station,
                        f"{actual:.2f}{unit}",
                        f"{std_value}{unit}",
                        f"{achievement*100:.2f}%",
                        f"{hours_ratio*100:.2f}%",
                        f"{contribution:.2f}%",
                        calc_process
                    ]

                    for col, val in enumerate(data_cells, 1):
                        cell = ws.cell(current_row, col, val)
                        self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                        if col == 7:  # 计算过程列
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                        # 达成率低于80%标红
                        if col == 4 and achievement < 0.8:
                            cell.fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
                            cell.font = Font(name=CHINESE_FONT, size=NORMAL_FONT_SIZE, color='FFFFFF')

                    ws.row_dimensions[current_row].height = 40
                    current_row += 1

                # 合计行 - 使用实际达成值
                ws.cell(current_row, 1, '合计')
                ws.merge_cells(f'B{current_row}:D{current_row}')
                ws.cell(current_row, 5, '-')  # 工时占比/产量占比列不显示加总
                ws.cell(current_row, 6, f'{actual_value:.2f}%')  # 只显示加权贡献
                ws.cell(current_row, 7, f'{weighted_score:.2f}分')

                for col in range(1, 8):
                    cell = ws.cell(current_row, col)
                    self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
                    cell.fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 25
                current_row += 2

            # 如果是返工率或报废率控制，显示各产品的计算
            if '控制' in indicator_name:
                # 添加公式说明行
                ws.merge_cells(f'A{current_row}:G{current_row}')
                formula_text = '计算公式：加权贡献 = (不良数÷产量) × 产量占比'
                cell = ws.cell(current_row, 1, formula_text)
                self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE, color=COLOR_SECONDARY)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                # 表头
                headers = ['产品', '产量', '不良数', '不良率', '产量占比', '加权贡献', '计算过程详细说明']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(current_row, col, header)
                    self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 30
                current_row += 1

                # 按产品聚合
                product_data = {}
                for agg in self.aggregated_data:
                    prod = agg.get('product', '')
                    if prod not in product_data:
                        product_data[prod] = {'bad': 0, 'total': 0}

                    if '返工' in indicator_name:
                        product_data[prod]['bad'] += agg.get('total_rework', 0)
                    else:
                        product_data[prod]['bad'] += agg.get('total_scrap', 0)
                    product_data[prod]['total'] += agg.get('total_good', 0) + agg.get('total_rework', 0) + agg.get('total_scrap', 0)

                grand_total = sum(p['total'] for p in product_data.values())

                for prod, data in product_data.items():
                    if data['total'] > 0:
                        bad_rate = data['bad'] / data['total'] * 100
                        output_ratio = data['total'] / grand_total * 100 if grand_total > 0 else 0
                        contribution = bad_rate * output_ratio / 100

                        # 计算过程说明
                        calc_process = (f'步骤1: {data["bad"]}件 ÷ {data["total"]}件 = {bad_rate:.2f}%\n'
                                       f'步骤2: {bad_rate:.2f}% × {output_ratio:.2f}% = {contribution:.2f}%')

                        data_cells = [
                            prod,
                            data['total'],
                            data['bad'],
                            f"{bad_rate:.2f}%",
                            f"{output_ratio:.2f}%",
                            f"{contribution:.2f}%",
                            calc_process
                        ]

                        for col, val in enumerate(data_cells, 1):
                            cell = ws.cell(current_row, col, val)
                            self._set_cell_font(cell, size=NORMAL_FONT_SIZE)
                            if col == 7:  # 计算过程列
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border

                            # 不良率超标标红
                            if col == 4:
                                if ('返工' in indicator_name and bad_rate > 0.5) or \
                                   ('报废' in indicator_name and bad_rate > 0.1):
                                    cell.fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
                                    cell.font = Font(name=CHINESE_FONT, size=NORMAL_FONT_SIZE, color='FFFFFF')

                        ws.row_dimensions[current_row].height = 40
                        current_row += 1

                # 合计行 - 使用实际达成值
                ws.cell(current_row, 1, '合计')
                ws.merge_cells(f'B{current_row}:D{current_row}')
                ws.cell(current_row, 5, '-')  # 工时占比/产量占比列不显示加总
                ws.cell(current_row, 6, f'{actual_value:.2f}%')  # 只显示加权贡献
                ws.cell(current_row, 7, f'{weighted_score:.2f}分')

                for col in range(1, 8):
                    cell = ws.cell(current_row, col)
                    self._set_cell_font(cell, bold=True, size=NORMAL_FONT_SIZE)
                    cell.fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 25
                current_row += 2

        # 调整列宽 - 7列，移除H列
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50

    def generate_excel(self, output_path: str):
        """生成完整的Excel报告"""
        wb = Workbook()

        # 删除默认sheet
        wb.remove(wb.active)

        # 添加各个sheet
        self._add_summary_sheet(wb)
        self._add_raw_data_sheet(wb)
        self._add_aggregated_sheet(wb)
        self._add_kpi_detail_sheet(wb)

        # 保存文件
        wb.save(output_path)
        print(f"报告已生成: {output_path}")


def generate_console_report(report_data: Dict):
    """生成控制台文本报告"""
    emp_info = report_data.get('employee_info', {})
    kpi_results = report_data.get('kpi_results', [])
    total_score = report_data.get('total_score', 0)
    total_grade = report_data.get('total_grade', '丁')

    print("\n" + "="*70)
    print(f"员工{emp_info.get('employee_name', '')}（工号：{emp_info.get('employee_id', '')}）")
    print(f"{emp_info.get('month', '')} KPI绩效计算结果")
    print("="*70)

    print("\n【KPI指标明细】")
    print("-"*70)
    print(f"{'KPI指标':<20} {'实际值':>12} {'等级':>6} {'原始分':>8} {'权重':>8} {'加权得分':>10}")
    print("-"*70)

    for kpi in kpi_results:
        name = kpi.get('indicator_name', '')
        value = kpi.get('actual_value', 0)
        grade = kpi.get('grade', '')
        raw = kpi.get('raw_score', 0)
        weight = kpi.get('weight', 0)
        weighted = kpi.get('weighted_score', 0)

        value_str = f"{value:.2f}%"
        print(f"{name:<20} {value_str:>12} {grade:>6} {raw:>8} {weight*100:>7.0f}% {weighted:>10.2f}")

    print("-"*70)
    print(f"{'月度综合绩效得分':<20} {'':>12} {total_grade:>6} {'':>8} {'':>8} {total_score:>10.2f}")
    print("="*70)

    # 计算验证
    print("\n【计算验证】")
    calc_sum = sum(k.get('weighted_score', 0) for k in kpi_results)
    score_parts = [f"{k.get('weighted_score', 0):.2f}" for k in kpi_results]
    print(f"各指标加权得分之和: {' + '.join(score_parts)} = {calc_sum:.2f}")
    verify_status = '验证通过' if abs(calc_sum - total_score) < 0.01 else '验证失败'
    print(f"与总得分对比: {total_score:.2f} [{verify_status}]")

    print("\n【中间计算结果 - 按产品+工站聚合】")
    print("-"*90)
    print(f"{'产品':<15} {'工站':<12} {'工时(h)':>10} {'良品数':>10} {'返工数':>8} {'报废数':>8} {'产出(件/h)':>12} {'良品率':>10}")
    print("-"*90)

    for agg in report_data.get('aggregated_data', []):
        print(f"{agg.get('product', ''):<15} {agg.get('station', ''):<12} "
              f"{agg.get('total_hours', 0):>10.1f} {agg.get('total_good', 0):>10} "
              f"{agg.get('total_rework', 0):>8} {agg.get('total_scrap', 0):>8} "
              f"{agg.get('actual_output', 0):>12.0f} {agg.get('actual_quality_rate', 0)*100:>9.1f}%")

    print("="*90)
